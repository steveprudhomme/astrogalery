#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import json
import time
import shutil
from datetime import datetime, date
from pathlib import Path
import xml.etree.ElementTree as ET

import requests
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from astropy.io import fits
from astropy.wcs import WCS
from astropy.visualization import ZScaleInterval, ImageNormalize

# Optional (carte atlas) - Open Source, rendu local (Hipparcos + lignes de constellations Stellarium)
# - nécessite une connexion Internet seulement au *premier* lancement pour télécharger:
#   1) le catalogue Hipparcos (Skyfield cache) et 2) le fichier constellationship.fab (Stellarium)
try:
    from skyfield.api import load as sf_load
    from skyfield.data import hipparcos as sf_hipparcos
    from astropy.coordinates import SkyCoord, SkyOffsetFrame
    import astropy.units as u
    HAS_ATLAS = True
except Exception:
    HAS_ATLAS = False

# Réglages carte atlas (finder chart)
ATLAS_FOV_ARCMIN = float(os.environ.get("GNU_ASTRO_GALERY_ATLAS_FOV_ARCMIN", "240"))  # champ total en arcmin (ex: 240 = 4°)
ATLAS_MAG_LIMIT = float(os.environ.get("GNU_ASTRO_GALERY_ATLAS_MAG_LIMIT", "10"))    # limite de magnitude (plus grand = plus d'étoiles)


from PIL import Image
from openpyxl import load_workbook


# -------------------------
# Configuration
# -------------------------
BOOTSTRAP_CDN_CSS = "https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css"
BOOTSTRAP_CDN_JS  = "https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"

ASTRO_NOVA_API  = "https://nova.astrometry.net/api/"
ASTRO_NOVA_SITE = "https://nova.astrometry.net/"

SIMBAD_TAP = "https://simbad.cds.unistra.fr/simbad/sim-tap/sync"

ASTROMETRY_MODE = "latest_per_object"  # "latest_per_object" or "all"
CACHE_PATH = Path("cache") / "object_info.json"

# Catalogue Messier (XLSX) placé au même endroit que le script
MESSIER_XLSX_NAME = "Objets Messiers..xlsx"

# Catalogue d’objets divers (XLSX multi-onglets) placé au même endroit que le script
DIVERSE_XLSX_NAME = "objetsdivers.xlsx"

# Limite magnitude pour labels de carte (objets divers)
DIVERSE_LABEL_MAG_LIMIT_DEFAULT = 6.0

# Cache astrométrie (persistant, hors de /site)
ASTRO_CACHE_DIR = Path("cache") / "astrometry"
ASTRO_CACHE_INDEX = ASTRO_CACHE_DIR / "index.json"
STAR_CACHE_DIR = Path("cache") / "starcharts"
STAR_CACHE_INDEX = STAR_CACHE_DIR / "index.json"

# Données (Stellarium) pour dessiner les lignes de constellations (carte atlas)
STELLARIUM_DATA_DIR = Path("data") / "stellarium"
# Depuis 2025+, Stellarium utilise des skycultures au format JSON (index.json) plutôt que constellationship.fab.
# On récupère donc les lignes de constellations depuis le dépôt open-source stellarium-skycultures.
CONSTELLATION_INDEX_JSON = STELLARIUM_DATA_DIR / "western_index.json"
CONSTELLATION_INDEX_URL = "https://raw.githubusercontent.com/Stellarium/stellarium-skycultures/master/western/index.json"


BASE_URL = "https://example.com/seestar"
SITE_TITLE = "Galerie Seestar S50"
CREATOR_NAME = "Steve Prud’Homme"
IMAGE_LICENSE = "Creative Commons CC0 1.0"

LOCAL_OBJECT_DB = {
    "M 51": ("galaxie spirale", "spiral galaxy",
             ["galaxie", "galaxie spirale", "interaction"],
             ["galaxy", "spiral galaxy", "interaction"]),
    "M 77": ("galaxie spirale", "spiral galaxy",
             ["galaxie", "galaxie spirale"],
             ["galaxy", "spiral galaxy"]),
    "M 94": ("galaxie spirale", "spiral galaxy",
             ["galaxie", "galaxie spirale"],
             ["galaxy", "spiral galaxy"]),
    "IC 342": ("galaxie spirale", "spiral galaxy",
               ["galaxie", "galaxie spirale"],
               ["galaxy", "spiral galaxy"]),
    "JUPITER": ("planète", "planet", ["planète"], ["planet"]),
    "DENEBOLA": ("étoile", "star", ["étoile"], ["star"]),
}

SIMBAD_OTYPE_MAP = {
    "G":   (["galaxie"], ["galaxy"]),
    "GiG": (["galaxie", "interaction"], ["galaxy", "interaction"]),
    "GPair": (["paire de galaxies"], ["galaxy pair"]),
    "ClG": (["amas de galaxies"], ["galaxy cluster"]),
    "GlC": (["amas globulaire"], ["globular cluster"]),
    "OpC": (["amas ouvert"], ["open cluster"]),
    "PN":  (["nébuleuse planétaire"], ["planetary nebula"]),
    "HII": (["région HII", "nébuleuse"], ["HII region", "nebula"]),
    "SNR": (["reste de supernova"], ["supernova remnant"]),
    "Neb": (["nébuleuse"], ["nebula"]),
    "RfN": (["nébuleuse par réflexion"], ["reflection nebula"]),
    "DNe": (["nébuleuse sombre"], ["dark nebula"]),
    "Star": (["étoile"], ["star"]),
    "Pl": (["planète"], ["planet"]),
    "SyS": (["étoile binaire"], ["binary star"]),
}


# ------------------------------------------------------------
# Utils
# ------------------------------------------------------------
def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s_-]+", "-", text, flags=re.UNICODE)
    return text[:120] if text else "unknown"


def safe_text(x, default=""):
    s = "" if x is None else str(x)
    s = s.strip()
    return s if s else default


def parse_date(date_obs: str):
    s = safe_text(date_obs, "")
    if not s:
        return None
    candidates = [
        ("%Y-%m-%dT%H:%M:%S", s[:19]),
        ("%Y-%m-%d %H:%M:%S", s[:19]),
        ("%Y-%m-%dT%H:%M:%S.%f", s),
        ("%Y-%m-%d %H:%M:%S.%f", s),
        ("%Y-%m-%d", s[:10]),
    ]
    for fmt, val in candidates:
        try:
            return datetime.strptime(val, fmt)
        except Exception:
            pass
    return None


def uniq_preserve(seq):
    seen = set()
    out = []
    for x in seq:
        if not x:
            continue
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def html_escape(s: str) -> str:
    s = s or ""
    return (s.replace("&", "&amp;")
              .replace("<", "&lt;")
              .replace(">", "&gt;")
              .replace('"', "&quot;")
              .replace("'", "&#039;"))


def load_json(path: Path) -> dict:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_json(path: Path, data: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def file_fingerprint(p: Path) -> str:
    """
    Fingerprint rapide et stable (sans lire tout le fichier), pratique en local.
    """
    st = p.stat()
    return f"{st.st_size}-{st.st_mtime_ns}"


# ------------------------------------------------------------
# IMPORTANT: SIMBAD ident = nom du répertoire (exclure _sub / -sub)
# ------------------------------------------------------------
def is_sub_dirname(name: str) -> bool:
    n = (name or "").strip().lower()
    return n.endswith("_sub") or n.endswith("-sub")


def simbad_ident_from_dir(obs_dir: Path) -> str:
    name = obs_dir.name.strip()
    if is_sub_dirname(name) and obs_dir.parent:
        name = obs_dir.parent.name.strip()
    return name


# ------------------------------------------------------------
# Cache SIMBAD
# ------------------------------------------------------------
def load_cache(cache_path: Path) -> dict:
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache_path: Path, cache: dict):
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


# ------------------------------------------------------------
# Catalogue Messier (XLSX)
# ------------------------------------------------------------
def normalize_messier_id(text: str) -> str | None:
    """
    Accepte: 'M1', 'M 1', 'm 001' -> 'M 1'
    """
    s = (text or "").strip().upper()
    m = re.search(r"\bM\s*0*([0-9]{1,3})\b", s)
    if not m:
        return None
    return f"M {int(m.group(1))}"


def parse_mag_cell(val) -> float | None:
    """
    Certaines magnitudes dans Excel peuvent être lues comme date.
    On reconstruit alors "jour.mois" (ex 2025-04-08 -> 8.4).
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    if isinstance(val, (datetime, date)):
        d = val.day
        m = val.month
        try:
            return round(d + (m / 10.0), 1)
        except Exception:
            return None
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        s2 = re.sub(r"[^\d./]", "", s)
        if "/" in s2:
            parts = s2.split("/")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                return round(int(parts[0]) + int(parts[1]) / 10.0, 1)
        return None


def extract_ngc_ic_from_name(name: str) -> str:
    """
    Extrait 'NGC 1952' ou 'IC 434' depuis 'NGC 1952 Nébuleuse du Crabe'
    """
    if not name:
        return ""
    m = re.search(r"\b(NGC|IC)\s*([0-9]{1,5})\b", str(name).upper())
    if not m:
        return ""
    return f"{m.group(1)} {int(m.group(2))}"


def load_messier_catalog(xlsx_path: Path) -> dict:
    """
    Retourne dict: 'M 1' -> champs (type, ngc, constellation, mag, taille, distance...)
    """
    if not xlsx_path.exists():
        return {}

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = {}
    for c in range(1, 40):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[str(v).strip()] = c

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    c_num = col("N°", "No", "N")
    c_name = col("Nom / NGC", "Nom/NGC", "Nom")
    c_type = col("Type")
    c_cons = col("Constellation")
    c_ra   = col("AD (h m)", "AD", "RA")
    c_dec  = col("Déc (° ')", "Dec", "DEC")
    c_mag  = col("Mag", "Magnitude")
    c_size = col("Taille", "Size")
    c_dist = col("Dist. (al)", "Dist", "Distance")

    if not c_num:
        return {}

    out = {}
    for r in range(2, ws.max_row + 1):
        n = ws.cell(r, c_num).value
        if n is None:
            continue
        try:
            n_int = int(float(n))
        except Exception:
            continue

        key = f"M {n_int}"

        name = ws.cell(r, c_name).value if c_name else ""
        otype = ws.cell(r, c_type).value if c_type else ""
        cons = ws.cell(r, c_cons).value if c_cons else ""
        ra = ws.cell(r, c_ra).value if c_ra else ""
        dec = ws.cell(r, c_dec).value if c_dec else ""
        mag_val = ws.cell(r, c_mag).value if c_mag else None
        size = ws.cell(r, c_size).value if c_size else ""
        dist = ws.cell(r, c_dist).value if c_dist else None

        mag = parse_mag_cell(mag_val)
        ngc_id = extract_ngc_ic_from_name(str(name) if name else "")

        try:
            dist_ly = float(dist) if dist is not None else None
            if dist_ly is not None:
                dist_ly = int(round(dist_ly))
        except Exception:
            dist_ly = None

        out[key.upper()] = {
            "m": n_int,
            "ngc_name": safe_text(name, ""),
            "ngc_id": safe_text(ngc_id, ""),
            "type": safe_text(otype, ""),
            "constellation": safe_text(cons, ""),
            "ra_hm": safe_text(ra, ""),
            "dec_dm": safe_text(dec, ""),
            "mag": mag,
            "size": safe_text(size, ""),
            "distance_ly": dist_ly,
        }
    return out



def _parse_ra_dec_to_deg(ra_raw, dec_raw):
    """Parse RA/Dec strings from the diverse XLSX into degrees (ICRS)."""
    try:
        from astropy.coordinates import SkyCoord
        import astropy.units as u
    except Exception:
        return None, None

    if ra_raw is None or dec_raw is None:
        return None, None
    ra = str(ra_raw).strip()
    dec = str(dec_raw).strip()
    if not ra or not dec:
        return None, None

    # Normalize unicode symbols commonly found in astro spreadsheets
    dec = (dec
           .replace("°", "d")
           .replace("º", "d")
           .replace("’", "m")
           .replace("′", "m")
           .replace("'", "m")
           .replace("″", "s")
           .replace("”", "s")
           .replace('"', "s")
           .replace(" ", ""))
    ra = ra.replace(" ", "")

    def _norm_hms(s: str) -> str:
        """Normalize RA HMS strings so seconds/minutes are within range (fixes 60s)."""
        s0 = s
        ss = s.strip()
        # Accept forms like 12h34m60s, 12:34:60, 12h34m60.0
        m = re.match(r'^([0-9]{1,3})(?:h|:)([0-9]{1,2})(?:m|:)([0-9]{1,2}(?:\.[0-9]+)?)s?$', ss)
        if not m:
            return s0
        h = int(m.group(1))
        mi = int(m.group(2))
        sec = float(m.group(3))
        # carry seconds -> minutes
        if sec >= 60.0:
            addm = int(sec // 60.0)
            sec = sec - 60.0 * addm
            mi += addm
        # carry minutes -> hours
        if mi >= 60:
            addh = mi // 60
            mi = mi % 60
            h += addh
        # Wrap within 0-23h
        h = h % 24
        return f"{h}h{mi}m{sec:.3f}s"

    def _norm_dms(s: str) -> str:
        """Normalize Dec DMS strings so seconds/minutes are within range (fixes 60s)."""
        s0 = s
        ss = s.strip()
        m = re.match(r'^([+-]?[0-9]{1,3})(?:d|:)([0-9]{1,2})(?:m|:)([0-9]{1,2}(?:\.[0-9]+)?)s?$', ss)
        if not m:
            return s0
        deg = int(m.group(1))
        sign = -1 if str(m.group(1)).startswith('-') else 1
        deg_abs = abs(deg)
        mi = int(m.group(2))
        sec = float(m.group(3))
        if sec >= 60.0:
            addm = int(sec // 60.0)
            sec = sec - 60.0 * addm
            mi += addm
        if mi >= 60:
            addd = mi // 60
            mi = mi % 60
            deg_abs += addd
        return f"{sign*deg_abs}d{mi}m{sec:.3f}s"

    # Fix occasional 60s in spreadsheets (avoids astropy IllegalSecondWarning)
    ra = _norm_hms(ra)
    dec = _norm_dms(dec)
    # Some sheets use "0h40.4" (decimal minutes) – SkyCoord handles it.
    try:
        c = SkyCoord(ra, dec, frame="icrs")
    except Exception:
        try:
            c = SkyCoord(ra=ra, dec=dec, unit=(u.hourangle, u.deg), frame="icrs")
        except Exception:
            return None, None
    return float(c.ra.deg), float(c.dec.deg)


def load_diverse_catalog(xlsx_path: Path) -> list[dict]:
    """
    Charge 'objetsdivers.xlsx' qui contient plusieurs onglets.
    Retourne une liste d'objets avec au minimum:
    - name, ra_deg, dec_deg, mag (float | None), sheet
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = []

    def _to_float(v):
        try:
            if v is None:
                return None
            s = str(v).strip()
            if not s:
                return None
            return float(s)
        except Exception:
            return None

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Find header row: expect 'Name', 'RA', 'Dec' at minimum
        header_row = None
        header = []
        for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            row_s = [str(c).strip() if c is not None else "" for c in row]
            joined = "|".join(row_s).lower()
            if "name" in joined and "ra" in joined and "dec" in joined:
                header_row = ridx
                header = row_s
                break
        if not header_row:
            continue

        # Map columns
        col = {h.strip().lower(): idx for idx, h in enumerate(header) if h and h.strip()}
        if "name" not in col or "ra" not in col or "dec" not in col:
            continue

        # Magnitude columns differ by sheet
        mag_key = None
        if "mag" in col:
            mag_key = "mag"
        elif "min mag" in col and "max mag" in col:
            mag_key = "minmax"
        elif "min mag" in col:
            mag_key = "min mag"

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row:
                continue
            name = row[col["name"]] if col["name"] < len(row) else None
            ra_raw = row[col["ra"]] if col["ra"] < len(row) else None
            dec_raw = row[col["dec"]] if col["dec"] < len(row) else None
            if name is None or ra_raw is None or dec_raw is None:
                continue
            name = str(name).strip()
            if not name:
                continue

            ra_deg, dec_deg = _parse_ra_dec_to_deg(ra_raw, dec_raw)
            if ra_deg is None or dec_deg is None:
                continue

            mag = None
            if mag_key == "mag":
                mag = _to_float(row[col["mag"]] if col["mag"] < len(row) else None)
            elif mag_key == "minmax":
                mn = _to_float(row[col["min mag"]] if col["min mag"] < len(row) else None)
                mx = _to_float(row[col["max mag"]] if col["max mag"] < len(row) else None)
                # take min magnitude if present
                mag = mn if mn is not None else mx
            elif mag_key == "min mag":
                mag = _to_float(row[col["min mag"]] if col["min mag"] < len(row) else None)

            out.append({
                "name": name,
                "ra_deg": ra_deg,
                "dec_deg": dec_deg,
                "mag": mag,
                "sheet": sheet,
            })

    return out


def find_diverse_xlsx(script_dir: Path, cwd: Path) -> Path | None:
    candidates = [
        script_dir / DIVERSE_XLSX_NAME,
        cwd / DIVERSE_XLSX_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    return None
def find_messier_xlsx(script_dir: Path, cwd: Path) -> Path | None:
    p1 = script_dir / MESSIER_XLSX_NAME
    if p1.exists():
        return p1
    p2 = cwd / MESSIER_XLSX_NAME
    if p2.exists():
        return p2
    for base in (script_dir, cwd):
        for p in base.glob("*.xlsx"):
            if "messier" in p.name.lower():
                return p
    return None


# ------------------------------------------------------------
# Catalogue / classification
# ------------------------------------------------------------
def normalize_catalog_id(name: str) -> str | None:
    s = (name or "").strip().upper()
    if s in ("JUPITER", "DENEBOLA"):
        return s

    m = re.search(r"\bM\s*([0-9]{1,3})\b", s)
    if m:
        return f"M {int(m.group(1))}"

    m = re.search(r"\bNGC\s*([0-9]{1,5})\b", s)
    if m:
        return f"NGC {int(m.group(1))}"

    m = re.search(r"\bIC\s*([0-9]{1,5})\b", s)
    if m:
        return f"IC {int(m.group(1))}"

    m = re.search(r"\bSH2[-\s]*([0-9]{1,4})\b", s)
    if m:
        return f"Sh2-{int(m.group(1))}"

    return None


def infer_catalog(object_name: str):
    up = (object_name or "").strip().upper()
    if re.match(r"^M\s*\d+", up):
        return "Messier"
    if up.startswith("NGC"):
        return "NGC"
    if up.startswith("IC"):
        return "IC"
    if up.startswith("SH2") or "SH2-" in up or "SH 2" in up:
        return "Sharpless"
    return "Other"


def infer_object_type_basic(object_name: str):
    up = (object_name or "").strip().upper()
    if up in ("JUPITER", "SATURN", "MARS", "VENUS"):
        return "Planet"
    if up in ("DENEBOLA", "ALTAIR", "VEGA", "DENEB"):
        return "Star"
    return "Other"


# ------------------------------------------------------------
# SIMBAD (TAP)
# ------------------------------------------------------------
def _simbad_query_basic(ident: str) -> dict | None:
    adql = f"""
    SELECT TOP 1 b.main_id, b.otype, b.otype_txt
    FROM basic AS b
    JOIN ident AS i ON i.oidref = b.oid
    WHERE i.id = '{ident.replace("'", "''")}'
    """
    r = requests.post(
        SIMBAD_TAP,
        data={"request": "doQuery", "lang": "adql", "format": "json", "query": adql})
    r.raise_for_status()
    data = r.json()
    rows = data.get("data", [])
    if not rows:
        return None
    fields = [f["name"] for f in data.get("fields", [])]
    row = dict(zip(fields, rows[0]))
    return {
        "main_id": row.get("main_id") or "",
        "otype": row.get("otype") or "",
        "otype_txt": row.get("otype_txt") or "",
    }



def _simbad_cone_basic(ra_deg: float, dec_deg: float, radius_deg: float, max_rows: int = 200) -> list[dict]:
    """Cone search SIMBAD TAP (table basic) autour d'un centre ICRS (RA/DEC en degrés).

    Retourne une liste de dicts: {main_id, ra, dec, otype, otype_txt}
    NOTE: on reste volontairement sur la table 'basic' (léger et stable).
    """
    try:
        r = float(radius_deg)
    except Exception:
        r = 2.0
    r = max(0.01, min(r, 10.0))

    adql = f"""
    SELECT TOP {int(max_rows)} b.main_id, b.ra, b.dec, b.otype, b.otype_txt
    FROM basic AS b
    WHERE 1=CONTAINS(
        POINT('ICRS', b.ra, b.dec),
        CIRCLE('ICRS', {ra_deg:.8f}, {dec_deg:.8f}, {r:.8f})
    )
    """
    resp = requests.post(
        SIMBAD_TAP,
        data={"request": "doQuery", "lang": "adql", "format": "json", "query": adql},
        timeout=60,
    )
    resp.raise_for_status()
    data = resp.json()
    out = []
    for row in data.get("data", []) or []:
        # colonne order: main_id, ra, dec, otype, otype_txt
        if not row or len(row) < 5:
            continue
        main_id = str(row[0] or "").strip()
        try:
            ra = float(row[1])
            dec = float(row[2])
        except Exception:
            continue
        otype = str(row[3] or "").strip()
        otype_txt = str(row[4] or "").strip()
        out.append({"main_id": main_id, "ra": ra, "dec": dec, "otype": otype, "otype_txt": otype_txt})
    return out


# classement "pertinence" pour les labels de carte
_FAMOUS_STAR_NAMES = {
    # très visibles à l'œil nu / noms très connus
    "SIRIUS", "CANOPUS", "ARCTURUS", "VEGA", "CAPELLA", "RIGEL", "PROCyon".upper(),
    "ACHERNAR", "BETELGEUSE", "HADAR", "ALTAIR", "ALDEBARAN", "ANTARES",
    "SPICA", "POLLUX", "FOMALHAUT", "DENEB", "REGULUS", "CASTOR",
    "ALPHA CENTAURI", "CENTAU".upper(), "CRUX".upper(), "ACRUX", "MIMOSA",
}

def _label_score(main_id: str, otype: str) -> tuple:
    """Retourne un tuple comparable: plus petit => plus important."""
    mid = (main_id or "").strip()
    up = mid.upper()
    # Messier d'abord
    if re.match(r"^M\s*\d+\b", up):
        # extraire numéro pour trier par proximité ensuite; ici on met catégorie
        return (0, 0, up)
    # Étoiles "fameuses" ensuite
    if any(name in up for name in _FAMOUS_STAR_NAMES):
        return (1, 0, up)
    # Autres étoiles (SIMBAD otype peut être '*' / 'Star' etc.)
    if (otype or "").strip() in {"*", "**", "V*", "PM*", "SB*", "EB*"} or "star" in (otype or "").lower():
        return (2, 0, up)
    # NGC / IC / Caldwell etc.
    if re.match(r"^(NGC|IC)\s*\d+\b", up):
        return (3, 0, up)
    # le reste
    return (4, 0, up)

def _clean_main_id_for_label(main_id: str) -> str:
    s = (main_id or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def enrich_tags(object_name_for_simbad: str, cache: dict) -> dict:
    ident = normalize_catalog_id(object_name_for_simbad) or object_name_for_simbad.strip()
    ident_key = ident.upper()

    if ident_key in LOCAL_OBJECT_DB:
        t_fr, t_en, tags_fr, tags_en = LOCAL_OBJECT_DB[ident_key]
        return {
            "ident": ident,
            "main_id": ident,
            "otype": "",
            "otype_txt": f"{t_en} / {t_fr}",
            "tags_fr": tags_fr,
            "tags_en": tags_en,
            "source": "local"
        }

    if ident_key in cache:
        return cache[ident_key]

    try:
        info = _simbad_query_basic(ident)
    except Exception as e:
        out = {"ident": ident, "main_id": "", "otype": "", "otype_txt": "", "tags_fr": [], "tags_en": [], "source": f"simbad_error:{e}"}
        cache[ident_key] = out
        return out

    if not info:
        out = {"ident": ident, "main_id": "", "otype": "", "otype_txt": "", "tags_fr": [], "tags_en": [], "source": "simbad_not_found"}
        cache[ident_key] = out
        return out

    otype = info.get("otype", "")
    otype_txt = info.get("otype_txt", "")

    tags_fr, tags_en = SIMBAD_OTYPE_MAP.get(otype, ([], []))
    if not tags_fr and otype_txt:
        tags_fr = [otype_txt.lower()]
    if not tags_en and otype_txt:
        tags_en = [otype_txt.lower()]

    out = {
        "ident": ident,
        "main_id": info.get("main_id", ""),
        "otype": otype,
        "otype_txt": otype_txt,
        "tags_fr": tags_fr,
        "tags_en": tags_en,
        "source": "simbad"
    }
    cache[ident_key] = out
    return out


# ------------------------------------------------------------
# FITS / metadata
# ------------------------------------------------------------
def extract_fits_metadata(fits_path: Path) -> dict:
    try:
        with fits.open(fits_path, ignore_missing_simple=True) as hdul:
            h = hdul[0].header
            return {
                "object": safe_text(h.get("OBJECT"), "Unknown Object"),
                "date_obs": safe_text(h.get("DATE-OBS"), ""),
                "exptime": h.get("EXPTIME", 0),
                "filter": safe_text(h.get("FILTER"), "None"),
                "telescope": safe_text(h.get("TELESCOP"), "Unknown Telescope"),
                "instrument": safe_text(h.get("INSTRUME"), "Unknown Instrument"),
                "observer": safe_text(h.get("OBSERVER"), "Unknown Observer"),
                "ra": safe_text(h.get("RA"), ""),
                "dec": safe_text(h.get("DEC"), ""),
            }
    except Exception as e:
        print(f"[WARN] FITS illisible: {fits_path} ({e})")
        return {}


def find_stacked_fits_in_dir(obs_dir: Path) -> Path | None:
    for p in obs_dir.rglob("Stacked*.fit*"):
        return p
    return None


def estimate_scale_arcsec_per_pix(fits_path: Path):
    try:
        with fits.open(fits_path, ignore_missing_simple=True) as hdul:
            h = hdul[0].header
        focal = h.get("FOCALLEN") or h.get("FOCAL") or h.get("FOCAL_LENGTH") or h.get("TELFOCAL")
        pix_um = h.get("XPIXSZ") or h.get("PIXSIZE") or h.get("PIXELSIZE") or h.get("PIX_SIZE") or h.get("XPIXELSZ")
        if focal is not None and pix_um is not None:
            focal = float(focal)     # mm
            pix_um = float(pix_um)   # microns
            if focal > 0 and pix_um > 0:
                return 206.265 * (pix_um / focal)
    except Exception:
        pass
    return None


# ------------------------------------------------------------
# Image read: FITS robust
# ------------------------------------------------------------
def _to_2d_array(arr: np.ndarray) -> np.ndarray | None:
    if arr is None:
        return None
    arr = np.array(arr)
    arr = np.squeeze(arr)

    if arr.ndim == 2:
        return arr
    if arr.ndim == 3:
        return arr[0] if arr.shape[0] <= 20 else arr[-1]
    if arr.ndim >= 4:
        while arr.ndim > 2:
            arr = arr[0]
        return arr if arr.ndim == 2 else None
    return None


def read_best_image_from_fits(fits_path: Path) -> np.ndarray | None:
    try:
        with fits.open(fits_path, ignore_missing_simple=True) as hdul:
            for hdu in hdul:
                data = getattr(hdu, "data", None)
                if data is None:
                    continue
                img = _to_2d_array(data)
                if img is not None and img.size > 0:
                    return img
    except Exception as e:
        print(f"[WARN] Lecture FITS échouée {fits_path.name}: {e}")
        return None
    return None


def read_image_from_jpg(jpg_path: Path) -> np.ndarray | None:
    try:
        im = Image.open(jpg_path).convert("L")
        return np.array(im)
    except Exception as e:
        print(f"[WARN] Lecture JPG échouée {jpg_path.name}: {e}")
        return None


# ------------------------------------------------------------
# JPG discovery (exclude *_sub/*-sub and *_thn.jpg)
# ------------------------------------------------------------
def is_in_sub_folder(path: Path) -> bool:
    for part in path.parts:
        p = str(part).lower()
        if p.endswith("_sub") or p.endswith("-sub"):
            return True
    return False


def is_thumbnail_file(path: Path) -> bool:
    return path.name.lower().endswith("_thn.jpg")


def find_final_jpgs(root_dir: Path):
    results = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        d = Path(dirpath)

        if d.name in ("site", ".git", "__pycache__", ".venv", "venv", "cache"):
            dirnames[:] = []
            continue

        if d.name.lower().endswith("_sub") or d.name.lower().endswith("-sub"):
            dirnames[:] = []
            continue

        dirnames[:] = [x for x in dirnames if not (x.lower().endswith("_sub") or x.lower().endswith("-sub"))]

        for fn in filenames:
            if not fn.lower().endswith(".jpg"):
                continue
            p = d / fn
            if is_in_sub_folder(p):
                continue
            if is_thumbnail_file(p):
                continue
            results.append(p)

    return results


# ------------------------------------------------------------
# Nova helpers
# ------------------------------------------------------------
def _json_or_raise(r: requests.Response, context: str) -> dict:
    r.raise_for_status()
    try:
        return r.json()
    except Exception:
        raise RuntimeError(
            f"{context}: réponse non-JSON. Content-Type={(r.headers.get('Content-Type') or '')}. Début: {r.text[:200]!r}"
        )


def nova_login(api_key: str) -> str:
    payload = {"apikey": api_key}
    r = requests.post(
        ASTRO_NOVA_API + "login",
        data={"request-json": json.dumps(payload)},
        headers={"Accept": "application/json"})
    data = _json_or_raise(r, "Login Nova échoué")
    if data.get("status") != "success":
        raise RuntimeError(f"Login Nova échoué: {data}")
    return data["session"]


def nova_upload_fits(session: str, fits_path: Path, scale_arcsec_per_pix=None):
    upload_kwargs = {
        "session": session,
        "allow_commercial_use": "d",
        "allow_modifications": "d",
        "publicly_visible": "n",
    }
    if scale_arcsec_per_pix:
        upload_kwargs.update({
            "scale_units": "arcsecperpix",
            "scale_est": str(scale_arcsec_per_pix),
            "scale_err": "0.25",
        })

    files = {"file": open(fits_path, "rb")}
    try:
        r = requests.post(
            ASTRO_NOVA_API + "upload",
            data={"request-json": json.dumps(upload_kwargs)},
            files=files,
            headers={"Accept": "application/json"})
        data = _json_or_raise(r, "Upload Nova échoué")
        if data.get("status") != "success":
            raise RuntimeError(f"Upload Nova échoué: {data}")
        return data["subid"]
    finally:
        files["file"].close()


def nova_poll_submission(subid: int, wait_s: int = 5, timeout_s: int = 600):
    t0 = time.time()
    while True:
        r = requests.get(ASTRO_NOVA_API + f"submissions/{subid}")
        data = _json_or_raise(r, f"Submission Nova échouée (subid={subid})")
        jobs = data.get("jobs") or []
        job_ids = [j for j in jobs if isinstance(j, int)]
        if job_ids:
            return job_ids[0]
        if time.time() - t0 > timeout_s:
            raise TimeoutError(f"Timeout: aucun job pour subid={subid} après {timeout_s}s")
        time.sleep(wait_s)


def nova_poll_job_solved(jobid: int, wait_s: int = 5, timeout_s: int = 900):
    t0 = time.time()
    while True:
        r = requests.get(ASTRO_NOVA_API + f"jobs/{jobid}")
        data = _json_or_raise(r, f"Job Nova échoué (jobid={jobid})")
        st = data.get("status")
        if st == "success":
            return True
        if st in ("failure", "error"):
            return False
        if time.time() - t0 > timeout_s:
            raise TimeoutError(f"Timeout: jobid={jobid} non résolu après {timeout_s}s")
        time.sleep(wait_s)


def looks_like_fits_bytes(b: bytes) -> bool:
    head = b[:80]
    try:
        s = head.decode("ascii", errors="ignore")
    except Exception:
        return False
    return "SIMPLE" in s


def download_binary(url: str, timeout: int = 300) -> tuple[bytes, str]:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content, (r.headers.get("Content-Type") or "")


def nova_download_wcs_header_only(jobid: int, out_fits_path: Path) -> bool:
    out_fits_path.parent.mkdir(parents=True, exist_ok=True)
    url = ASTRO_NOVA_SITE + f"wcs_file/{jobid}"
    b, ct = download_binary(url)
    if not looks_like_fits_bytes(b):
        snippet = b[:300].decode("utf-8", errors="ignore")
        print(f"\n[WARN] wcs_file non-FITS jobid={jobid} (Content-Type={ct}). Début: {snippet!r}")
        return False
    out_fits_path.write_bytes(b)
    print(f"\n[OK] Téléchargé WCS header-only jobid={jobid} -> {out_fits_path.name} (Content-Type={ct})")
    return True


def load_wcs_header_only(wcs_fits_path: Path) -> fits.Header | None:
    try:
        with fits.open(wcs_fits_path, ignore_missing_simple=True) as hdul:
            return hdul[0].header
    except Exception as e:
        print(f"[WARN] Lecture WCS header-only échouée {wcs_fits_path.name}: {e}")
        return None


# ------------------------------------------------------------
# ASTROMETRY PNG (image + WCS header-only)
# ------------------------------------------------------------
def make_astrometry_png_from_image_and_wcs(
    image_array_2d: np.ndarray,
    wcs_header: fits.Header,
    out_png: Path,
    title: str = ""
) -> bool:
    try:
        h = wcs_header.copy()
        ny, nx = image_array_2d.shape
        if h.get("NAXIS", None) is None:
            h["NAXIS"] = 2
        if h.get("NAXIS1", None) is None:
            h["NAXIS1"] = nx
        if h.get("NAXIS2", None) is None:
            h["NAXIS2"] = ny

        wcs = WCS(h, naxis=2)
        norm = ImageNormalize(image_array_2d, interval=ZScaleInterval())

        fig = plt.figure(figsize=(6, 9), dpi=160)
        ax = fig.add_subplot(111, projection=wcs)

        ax.imshow(image_array_2d, origin="lower", norm=norm)
        ax.grid(color="white", alpha=0.35, linestyle="-", linewidth=0.6)
        ax.set_xlabel("RA")
        ax.set_ylabel("DEC")

        if title:
            ax.set_title(title, fontsize=11)

        cx, cy = nx / 2, ny / 2
        radius_pix = min(nx, ny) * 0.33
        ax.add_patch(plt.Circle((cx, cy), radius_pix, fill=False, lw=1.2, alpha=0.85))

        out_png.parent.mkdir(parents=True, exist_ok=True)
        fig.tight_layout()
        fig.savefig(out_png)
        plt.close(fig)

        print(f"[OK] PNG écrit (image+WCS): {out_png}")
        return True
    except Exception as e:
        print(f"\n[WARN] PNG astrométrie impossible: {e}")
        return False


# ------------------------------------------------------------
# STAR CHART (Finder chart) + cache persistant
# ------------------------------------------------------------
def load_star_cache() -> dict:
    return load_json(STAR_CACHE_INDEX)

def save_star_cache(cache: dict):
    save_json(STAR_CACHE_INDEX, cache)

def wcs_center_from_header(h: fits.Header) -> tuple[float | None, float | None]:
    try:
        ra = h.get("CRVAL1", None)
        dec = h.get("CRVAL2", None)
        ra = float(ra) if ra is not None else None
        dec = float(dec) if dec is not None else None
        return ra, dec
    except Exception:
        return None, None

def _ensure_constellation_index_json() -> bool:
    """
    Assure la présence du fichier 'western/index.json' (format skyculture JSON).
    Source: dépôt open-source Stellarium/stellarium-skycultures.
    Cache local: {STELLARIUM_DATA_DIR}
    Sortie: {CONSTELLATION_INDEX_JSON}
    """
    try:
        if CONSTELLATION_INDEX_JSON.exists() and CONSTELLATION_INDEX_JSON.stat().st_size > 1024:
            return True

        STELLARIUM_DATA_DIR.mkdir(parents=True, exist_ok=True)

        import requests
        r = requests.get(CONSTELLATION_INDEX_URL, timeout=60)
        r.raise_for_status()
        # L'index est du JSON
        data = r.content
        if not data or len(data) < 1024:
            raise IOError("index.json trop petit / invalide.")
        CONSTELLATION_INDEX_JSON.write_bytes(data)
        return True

    except Exception as e:
        print(f"[WARN] Carte atlas: téléchargement index.json (constellations) impossible: {e}")
        return False


_HIP_DF = None

_CONSTELLATION_LINES = None


def _load_hipparcos_df():
    """Charge le dataframe Hipparcos via Skyfield (cache disque + cache mémoire)."""
    global _HIP_DF
    if _HIP_DF is not None:
        return _HIP_DF
    try:
        # Skyfield gère le cache disque automatiquement (dans son répertoire de cache)
        with sf_load.open(sf_hipparcos.URL) as f:
            df = sf_hipparcos.load_dataframe(f)
        # df index = HIP
        _HIP_DF = df
        return df
    except Exception as e:
        print(f"[WARN] Carte atlas: chargement Hipparcos impossible: {e}")
        return None


def _load_constellation_lines():
    """
    Charge les lignes de constellations depuis 'western/index.json' (Stellarium skyculture JSON).

    Retour: list[tuple[str, list[tuple[int,int]]]]
      - str: abréviation IAU (ex: 'Aql') si présente, sinon identifiant de constellation
      - segments: liste de paires (HIP_a, HIP_b)
    """
    global _CONSTELLATION_LINES
    if _CONSTELLATION_LINES is not None:
        return _CONSTELLATION_LINES

    if not _ensure_constellation_index_json():
        _CONSTELLATION_LINES = []
        return _CONSTELLATION_LINES

    import json

    try:
        data = json.loads(CONSTELLATION_INDEX_JSON.read_text(encoding="utf-8", errors="ignore"))
        consts = data.get("constellations", [])
        out = []
        for c in consts:
            label = c.get("iau") or c.get("id") or "CON"
            segs = []
            for path in c.get("lines", []) or []:
                # path: [hip, hip, hip] ou ["thin"/"bold", hip, hip, ...]
                if not path:
                    continue
                if isinstance(path[0], str):
                    stars = path[1:]
                else:
                    stars = path
                # segments successifs
                for a, b in zip(stars, stars[1:]):
                    try:
                        ia = int(a); ib = int(b)
                        segs.append((ia, ib))
                    except Exception:
                        continue
            if segs:
                out.append((label, segs))

        _CONSTELLATION_LINES = out
        return _CONSTELLATION_LINES

    except Exception as e:
        print(f"[WARN] Carte atlas: lecture/parsing index.json impossible: {e}")
        _CONSTELLATION_LINES = []
        return _CONSTELLATION_LINES


def make_finder_chart_png(ra_deg: float, dec_deg: float, out_png: Path, fov_arcmin: float | None = None, inner_fov_arcmin: float = 30.0, title: str = "", diverse_catalog: list[dict] | None = None, diverse_mag_limit: float = DIVERSE_LABEL_MAG_LIMIT_DEFAULT) -> bool:
    """
    Génère une carte 'atlas' (grille RA/Dec, étoiles, lignes de constellations) centrée sur (RA,DEC).
    - rendu local open source (matplotlib + astropy + skyfield)
    - nécessite une connexion Internet seulement au *premier* lancement (cache Skyfield + Stellarium)
    """
    if not HAS_ATLAS:
        print("[WARN] Carte atlas: dépendances manquantes (pip install skyfield)")
        return False

    df = _load_hipparcos_df()
    if df is None:
        return False

    cons = _load_constellation_lines()

    try:
        center = SkyCoord(ra=ra_deg * u.deg, dec=dec_deg * u.deg, frame="icrs")
        frame = SkyOffsetFrame(origin=center)
        if fov_arcmin is None:
            fov_arcmin = ATLAS_FOV_ARCMIN
        radius_deg = (fov_arcmin / 60.0) / 2.0
        margin_deg = radius_deg * 1.35

        # Pré-sélection rapide par boîte (sur la sphère c'est approximatif mais OK pour réduire)
        ra = df["ra_degrees"].to_numpy()
        dec = df["dec_degrees"].to_numpy()
        mag = df["magnitude"].to_numpy()

        # distance angulaire précise avec astropy (sur subset)
        stars = SkyCoord(ra=ra * u.deg, dec=dec * u.deg, frame="icrs")
        sep = stars.separation(center).deg
        msk = (sep <= margin_deg) & (mag == mag)  # mag not NaN
        if msk.sum() == 0:
            print("[WARN] Carte atlas: aucune étoile Hipparcos dans le champ")
            return False

        stars_sel = stars[msk].transform_to(frame)
        mag_sel = mag[msk]

        x = stars_sel.lon.to(u.deg).value * 60.0  # arcmin
        y = stars_sel.lat.to(u.deg).value * 60.0

        # taille des points: magnitude -> taille
        # Hipparcos: mag plus petite = plus brillant
        mag_limit = ATLAS_MAG_LIMIT
        s = (np.clip((mag_limit - mag_sel + 1.0), 0.2, 6.0) ** 2) * 3.0

        out_png.parent.mkdir(parents=True, exist_ok=True)

        fig = plt.figure(figsize=(8, 8), dpi=160)
        ax = plt.gca()
        ax.scatter(x, y, s=s, alpha=0.85)

        # convention carte du ciel: Est à gauche
        ax.invert_xaxis()

        half = radius_deg * 60.0
        ax.set_xlim(half, -half)
        ax.set_ylim(-half, half)
        ax.set_aspect("equal", "box")

        # Grille (arcmin)
        ax.grid(True, linewidth=0.4, alpha=0.6)
        ax.set_xlabel("ΔRA cos(Dec) (arcmin)")
        ax.set_ylabel("ΔDec (arcmin)")


        # cercle FOV (Seestar) à l'intérieur de la carte (atlas)
        inner_half = max(1.0, float(inner_fov_arcmin) / 2.0)  # arcmin
        ax.add_patch(plt.Circle((0, 0), inner_half, fill=False, linewidth=1.2))
        ax.annotate(f"FOV {inner_fov_arcmin:.0f}'", xy=(inner_half, 0), xytext=(inner_half+3, 0), va="center")


        # Lignes de constellations (si dispo)
        # On trace seulement les segments dont les 2 étoiles tombent dans la marge
        if cons:
            # Accès rapide HIP -> (x,y) dans la même projection:
            # On recalcule pour les HIPs utiles en évitant de transformer tout le catalogue.
            # Dictionnaire hip->SkyCoord projeté
            # df index contient les HIP
            hip_xy = {}
            def get_xy(hip:int):
                if hip in hip_xy:
                    return hip_xy[hip]
                try:
                    row = df.loc[hip]
                    sc = SkyCoord(ra=float(row["ra_degrees"]) * u.deg, dec=float(row["dec_degrees"]) * u.deg, frame="icrs")
                    if sc.separation(center).deg > margin_deg:
                        hip_xy[hip] = None
                        return None
                    off = sc.transform_to(frame)
                    xx = off.lon.to(u.deg).value * 60.0
                    yy = off.lat.to(u.deg).value * 60.0
                    hip_xy[hip] = (xx, yy)
                    return (xx, yy)
                except Exception:
                    hip_xy[hip] = None
                    return None

            for abbr, segs in cons:
                pts_for_label = []
                for h1, h2 in segs:
                    p1 = get_xy(h1)
                    p2 = get_xy(h2)
                    if p1 is None or p2 is None:
                        continue
                    ax.plot([p1[0], p2[0]], [p1[1], p2[1]], linewidth=0.6, alpha=0.7)
                    pts_for_label.append(p1)
                    pts_for_label.append(p2)
                # label: moyenne si on a assez de points
                if pts_for_label:
                    xs = [p[0] for p in pts_for_label]
                    ys = [p[1] for p in pts_for_label]
                    cx = float(np.mean(xs))
                    cy = float(np.mean(ys))
                    if -half < cy < half and -half < cx < half:
                        ax.text(cx, cy, abbr, fontsize=8, alpha=0.8, ha="center", va="center")

        # marqueur cible
        ax.scatter([0], [0], s=80, marker="x")

        # label de l'objet au centre (cible)
        target_label = title.strip()
        if target_label:
            ax.text(0.0, -half * 0.92, target_label, fontsize=10, ha="center", va="top")

        # objets "pertinents" proches via SIMBAD (Messier d'abord, puis étoiles connues/lumineuses)
        try:
            nearby = _simbad_cone_basic(ra_deg, dec_deg, radius_deg=margin_deg, max_rows=300)
        except Exception as e:
            print(f"[WARN] Carte atlas: requête SIMBAD (cone) impossible: {e}")
            nearby = []

        # Projeter et choisir quelques labels sans trop d'encombrement
        if nearby or diverse_catalog:
            placed = []  # list of (x,y)
            max_labels = 10

            # --- Labels pertinents (atlas) ---
            # Règle: afficher uniquement
            #  1) Objets Messier (via SIMBAD cone) dans le champ
            #  2) Objets provenant de 'objetsdivers.xlsx' avec magnitude <= diverse_mag_limit dans le champ
            label_candidates = []

            # 1) Messier (depuis SIMBAD cone)
            for it in nearby:
                mid_raw = str(it.get("main_id","") or "").strip()
                if re.match(r"^M\s*\d+\b", mid_raw.upper()):
                    label_candidates.append(it)

            # 2) Objets divers (catalogue local)
            if diverse_catalog:
                try:
                    center_icrs = SkyCoord(ra=ra_deg * u.deg, dec=dec_deg * u.deg, frame="icrs")
                    half_deg = (fov_arcmin / 60.0) / 2.0
                    for ob in diverse_catalog:
                        mag = ob.get("mag", None)
                        if mag is None:
                            continue
                        try:
                            if float(mag) > float(diverse_mag_limit):
                                continue
                        except Exception:
                            continue
                        try:
                            c = SkyCoord(ra=float(ob["ra_deg"]) * u.deg, dec=float(ob["dec_deg"]) * u.deg, frame="icrs")
                            if float(c.separation(center_icrs).deg) <= (half_deg * 1.05):
                                label_candidates.append({
                                    "main_id": ob.get("name",""),
                                    "ra": float(ob["ra_deg"]),
                                    "dec": float(ob["dec_deg"]),
                                    "otype": "CAT",
                                    "otype_txt": f'{ob.get("sheet","")}',
                                    "mag": float(mag),
                                })
                        except Exception:
                            continue
                except Exception:
                    pass

            # trier: Messier d'abord, puis magnitude ascendante, puis distance au centre
            center_icrs = SkyCoord(ra=ra_deg * u.deg, dec=dec_deg * u.deg, frame="icrs")

            def _is_messier(it):
                return bool(re.match(r"^M\s*\d+\b", str(it.get("main_id","")).upper()))

            def _mag(it):
                try:
                    return float(it.get("mag", 99.0))
                except Exception:
                    return 99.0

            def _dist_deg(it):
                try:
                    c = SkyCoord(ra=float(it["ra"]) * u.deg, dec=float(it["dec"]) * u.deg, frame="icrs")
                    return float(c.separation(center_icrs).deg)
                except Exception:
                    return 999.0

            nearby_sorted = sorted(
                label_candidates,
                key=lambda it: (0 if _is_messier(it) else 1, _mag(it), _dist_deg(it))
            )
            for it in nearby_sorted:
                mid = _clean_main_id_for_label(it.get("main_id",""))
                if not mid:
                    continue
                # ne pas répéter le titre central si c'est le même
                if target_label and mid.upper() == target_label.upper():
                    continue

                try:
                    c = SkyCoord(
                        ra=float(it["ra"]) * u.deg,
                        dec=float(it["dec"]) * u.deg,
                        frame="icrs",
                    ).transform_to(frame)
                    # IMPORTANT: la carte est en **arcmin** (comme les étoiles Hipparcos plus haut)
                    # Donc on convertit lon/lat (degrés) -> arcmin pour positionner correctement les labels.
                    x = float(c.lon.to(u.deg).value * 60.0)
                    y = float(c.lat.to(u.deg).value * 60.0)
                except Exception:
                    continue

                if not (-half < x < half and -half < y < half):
                    continue

                # anti-collision simple
                ok = True
                # (x,y) en arcmin -> distance mini en arcmin
                min_sep = 6.0  # arcmin (évite que les noms se pile au centre)
                for (px, py) in placed:
                    if (x - px) ** 2 + (y - py) ** 2 < (min_sep ** 2):
                        ok = False
                        break
                if not ok:
                    continue

                ax.scatter([x], [y], s=18, marker="o", alpha=0.9)
                # placer le texte légèrement sous le symbole (en arcmin)
                ax.text(x, y - 2.0, mid, fontsize=7.5, ha="center", va="top", alpha=0.9)
                placed.append((x, y))

                if len(placed) >= max_labels:
                    break


        ttl = title.strip() or "Carte (atlas)"
        ax.set_title(ttl)

        fig.savefig(out_png, bbox_inches="tight")
        plt.close(fig)
        return True
    except Exception as e:
        print(f"[WARN] Carte atlas impossible: {e}")
        return False


def image_jsonld(item: dict, page_url: str) -> dict:
    add_props = []
    for k in ["ra", "dec", "exptime", "filter", "telescope", "instrument",
              "messier", "ngc", "constellation", "magnitude", "size", "distance_ly",
              "simbad_main_id", "simbad_otype", "simbad_otype_txt"]:
        v = item.get(k)
        if v not in (None, "", 0, "0"):
            add_props.append({"@type": "PropertyValue", "name": k.upper(), "value": str(v)})

    keywords = uniq_preserve(
        (item.get("keywords_fr", []) + item.get("keywords_en", []) + item.get("tags_fr", []) + item.get("tags_en", []))
    )

    return {
        "@context": "https://schema.org",
        "@type": "ImageObject",
        "name": item.get("objectName") or item.get("name"),
        "description": item.get("description", ""),
        "keywords": ", ".join(keywords),
        "contentUrl": item["contentUrlAbs"],
        "thumbnailUrl": item["thumbnailUrlAbs"],
        "dateCreated": item.get("dateCreatedISO", "") or item.get("dateCreated", ""),
        "about": {"@type": "Thing", "name": item.get("objectName", "")},
        "isPartOf": {"@type": "CollectionPage", "name": SITE_TITLE, "url": BASE_URL.rstrip("/") + "/index.html"},
        "creator": {"@type": "Person", "name": CREATOR_NAME},
        "license": IMAGE_LICENSE,
        "url": page_url,
        "additionalProperty": add_props
    }




def og_meta(title: str, description: str, image_url: str, page_url: str) -> str:
    """OpenGraph + Twitter meta tags (used for link previews)."""
    # Note: keep values already escaped by caller when needed.
    return f"""
<meta property="og:type" content="website">
<meta property="og:title" content="{title}">
<meta property="og:description" content="{description}">
<meta property="og:image" content="{image_url}">
<meta property="og:url" content="{page_url}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{title}">
<meta name="twitter:description" content="{description}">
<meta name="twitter:image" content="{image_url}">
"""

# ------------------------------------------------------------
# HTML builders
# ------------------------------------------------------------
def build_index_html(title: str, og_block: str) -> str:
    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(title)}</title>
  <meta name="description" content="Galerie astrophotographie — Seestar S50">
{og_block}
  <link rel="stylesheet" href="{BOOTSTRAP_CDN_CSS}">
  <link rel="stylesheet" href="assets/css/styles.css">
</head>
<body class="bg-body-tertiary">
<nav class="navbar navbar-expand-lg navbar-dark bg-dark">
  <div class="container">
    <a class="navbar-brand" href="index.html">{html_escape(title)}</a>
  </div>
</nav>

<main class="container py-4">
  <div class="row g-3 align-items-end">
    <div class="col-md-6">
      <label class="form-label">Recherche</label>
      <input id="searchInput" class="form-control" placeholder="M31, NGC7000, nebula, galaxy...">
    </div>
    <div class="col-md-3">
      <label class="form-label">Type</label>
      <select id="typeSelect" class="form-select">
        <option value="">Tous / All</option>
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">Catalogue</label>
      <select id="catalogSelect" class="form-select">
        <option value="">Tous / All</option>
      </select>
    </div>
  </div>

  <hr class="my-4">
  <div id="stats" class="text-muted small mb-3"></div>
  <div id="grid" class="row g-3"></div>
</main>

<script id="images-data" type="application/json"></script>

<script src="{BOOTSTRAP_CDN_JS}"></script>
<script src="assets/js/app.js"></script>
</body>
</html>
"""


def build_object_page_html(site_title: str, obj_name: str, jsonld_block: str, og_block: str, items: list) -> str:
    # Héro = plus récent (items[0] est trié ailleurs)
    hero = items[0]

    author = CREATOR_NAME
    license_name = IMAGE_LICENSE

    hero_img = hero["contentUrl"]
    hero_img_alt = hero.get("alt", obj_name)

    # --- Métadonnées de l'image ---
    image_meta_rows = [
        ("Nom de fichier", hero.get("name", "")),
        ("Date d’acquisition", hero.get("dateCreated", "")),
        ("Exposition (s)", hero.get("exptime", "")),
        ("Filtre", hero.get("filter", "")),
        ("Télescope", hero.get("telescope", "")),
        ("Instrument", hero.get("instrument", "")),
        ("RA", hero.get("ra", "")),
        ("DEC", hero.get("dec", "")),
        ("Catalogue", hero.get("catalog", "")),
        ("Type (UI)", hero.get("objectType", "")),
    ]
    image_meta_rows = [(k, v) for k, v in image_meta_rows if str(v).strip() not in ("", "None", "0")]

    meta_table = "\n".join(
        f"<tr><th scope='row' class='w-50'>{html_escape(k)}</th><td>{html_escape(str(v))}</td></tr>"
        for k, v in image_meta_rows
    ) or "<tr><td colspan='2' class='text-muted'>Aucune métadonnée disponible.</td></tr>"

    # --- Caractéristiques de l'objet (SIMBAD + Messier XLSX) ---
    obj_rows = [
        ("Identification (SIMBAD query)", hero.get("simbad_ident", "")),
        ("Main ID (SIMBAD)", hero.get("simbad_main_id", "")),
        ("OTYPE (SIMBAD)", hero.get("simbad_otype", "")),
        ("Type (SIMBAD)", hero.get("simbad_otype_txt", "")),
        ("Source (SIMBAD)", hero.get("simbad_source", "")),

        ("Messier", hero.get("messier", "")),
        ("NGC/IC", hero.get("ngc", "")),
        ("Constellation", hero.get("constellation", "")),
        ("Magnitude", hero.get("magnitude", "")),
        ("Taille", hero.get("size", "")),
        ("Distance (al)", hero.get("distance_ly", "")),
        ("Type (catalogue Messier)", hero.get("messier_type", "")),
    ]
    obj_rows = [(k, v) for k, v in obj_rows if str(v).strip() not in ("", "None", "0")]

    obj_table = "\n".join(
        f"<tr><th scope='row' class='w-50'>{html_escape(k)}</th><td>{html_escape(str(v))}</td></tr>"
        for k, v in obj_rows
    ) or "<tr><td colspan='2' class='text-muted'>Aucune caractéristique disponible.</td></tr>"

    # --- Tags (badges) ---
    tags = uniq_preserve((hero.get("tags_fr", []) + hero.get("tags_en", [])))
    tags_badges = " ".join(
        f"<span class='badge text-bg-secondary me-1 mb-1'>{html_escape(t)}</span>"
        for t in tags
    ) or "<span class='text-muted'>Aucun tag</span>"

    # --- Astrométrie (preview + modal) ---
    astro = hero.get("astrometryUrl", "")
    astro_block = ""
    astro_modal = ""
    if astro:
        astro_id = "astroModal"
        astro_block = f"""
        <div class="card shadow-sm">
          <div class="card-body">
            <div class="d-flex align-items-baseline justify-content-between">
              <div class="fw-semibold">Astrométrie</div>
              <span class="text-muted small">Cliquez pour agrandir</span>
            </div>
            <a href="#" data-bs-toggle="modal" data-bs-target="#{astro_id}">
              <img src="../{html_escape(astro)}" class="img-fluid rounded mt-2 astro-preview" alt="Astrométrie {html_escape(obj_name)}">
            </a>
          </div>
        </div>
        """

        astro_modal = f"""
<div class="modal fade" id="{astro_id}" tabindex="-1" aria-hidden="true">
  <div class="modal-dialog modal-xl modal-dialog-centered">
    <div class="modal-content">
      <div class="modal-header">
        <h5 class="modal-title">Astrométrie — {html_escape(obj_name)}</h5>
        <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Fermer"></button>
      </div>
      <div class="modal-body text-center">
        <img src="../{html_escape(astro)}" class="img-fluid rounded" alt="Astrométrie {html_escape(obj_name)}">
      </div>
    </div>
  </div>
</div>
"""

    # --- Autres images ---
    gallery_cards = []
    for it in items:
        gallery_cards.append(f"""
        <div class="col-md-4">
          <div class="card h-100 shadow-sm">
            <a href="../{html_escape(it['contentUrl'])}" target="_blank" rel="noopener">
              <img src="../{html_escape(it['thumbnailUrl'])}" class="card-img-top" alt="{html_escape(it.get('alt',''))}">
            </a>
            <div class="card-body">
              <div class="fw-semibold">{html_escape(it.get('objectName', it.get('name','')))}</div>
              <div class="text-muted small">{html_escape(it.get('dateCreated',''))}</div>
            </div>
          </div>
        </div>
        """)
    gallery_cards_html = "\n".join(gallery_cards)

    # --- Carte stellaire (finder chart) ---
    star = hero.get("starChartUrl", "")
    star_block = ""
    star_modal = ""
    if star:
        star_id = "starChartModal"
        star_block = f"""
        <div class="card shadow-sm mt-3">
          <div class="card-body">
            <div class="d-flex align-items-baseline justify-content-between">
              <div class="fw-semibold">Carte (atlas)</div>
              <span class="text-muted small">Cliquez pour agrandir</span>
            </div>
            <a href="#" data-bs-toggle="modal" data-bs-target="#{star_id}">
              <img src="../{html_escape(star)}" class="img-fluid mt-2 astro-preview" alt="Carte stellaire {html_escape(obj_name)}">
            </a>
          </div>
        </div>
        """
    
        star_modal = f"""
        <div class="modal fade" id="{star_id}" tabindex="-1" aria-hidden="true">
          <div class="modal-dialog modal-xl modal-dialog-centered">
            <div class="modal-content">
              <div class="modal-header">
                <h5 class="modal-title">Carte (atlas) — {html_escape(obj_name)}</h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
              </div>
              <div class="modal-body">
                <img src="../{html_escape(star)}" class="img-fluid" alt="Carte stellaire {html_escape(obj_name)}">
              </div>
            </div>
          </div>
        </div>
        """

    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html_escape(obj_name)} — {html_escape(site_title)}</title>
  <meta name="description" content="Astrophotographie : {html_escape(obj_name)}">
{og_block}
  <link rel="stylesheet" href="{BOOTSTRAP_CDN_CSS}">
  <link rel="stylesheet" href="../assets/css/styles.css">
  <script type="application/ld+json">
{jsonld_block}
  </script>
</head>
<body class="bg-body-tertiary">

<nav class="navbar navbar-expand-lg navbar-dark bg-dark">
  <div class="container">
    <a class="navbar-brand" href="../index.html">{html_escape(site_title)}</a>
    <span class="navbar-text text-white-50 ms-3">{html_escape(obj_name)}</span>
  </div>
</nav>

<main class="container py-4">
  <div class="d-flex flex-wrap align-items-end justify-content-between gap-2 mb-3">
    <div>
      <h1 class="h3 mb-1">{html_escape(obj_name)}</h1>
      <div class="text-muted small">
        Auteur : <span class="fw-semibold">{html_escape(author)}</span> • Licence : <span class="fw-semibold">{html_escape(license_name)}</span>
      </div>
    </div>
    <div class="text-end">
      <a class="btn btn-outline-secondary btn-sm" href="../index.html">← Retour</a>
      <a class="btn btn-primary btn-sm" href="../{html_escape(hero_img)}" target="_blank" rel="noopener">Ouvrir l’image</a>
    </div>
  </div>

  <!-- Hero -->
  <div class="card shadow-sm mb-4">
    <div class="row g-0">
      <div class="col-lg-8">
        <a href="../{html_escape(hero_img)}" target="_blank" rel="noopener">
          <img src="../{html_escape(hero_img)}" class="img-fluid w-100 object-hero" alt="{html_escape(hero_img_alt)}">
        </a>
      </div>
      <div class="col-lg-4">
        <div class="card-body">
          <div class="fw-semibold mb-2">Tags</div>
          <div class="mb-3">{tags_badges}</div>

          <div class="fw-semibold mb-2">Résumé</div>
          <div class="small text-muted">
            Catalogue : <span class="text-body">{html_escape(hero.get("catalog",""))}</span><br>
            Type : <span class="text-body">{html_escape(hero.get("objectType",""))}</span><br>
            Filtre : <span class="text-body">{html_escape(str(hero.get("filter","")))}</span>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- Two tables -->
  <div class="row g-4 mb-4">
    <div class="col-lg-6">
      <div class="card shadow-sm h-100">
        <div class="card-body">
          <h2 class="h5 mb-3">Métadonnées de l’image</h2>
          <div class="table-responsive">
            <table class="table table-sm align-middle">
              <tbody>
                {meta_table}
              </tbody>
            </table>
          </div>
        </div>
      </div>
    </div>

    <div class="col-lg-6">
      <div class="card shadow-sm h-100">
        <div class="card-body">
          <h2 class="h5 mb-3">Caractéristiques de l’objet</h2>
          <div class="table-responsive">
            <table class="table table-sm align-middle">
              <tbody>
                {obj_table}
              </tbody>
            </table>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- Astrometry -->
  <div class="mb-4">
    {astro_block}
        {star_block}
  </div>

  <!-- Other exposures -->
  <div class="d-flex align-items-baseline justify-content-between mb-2">
    <h2 class="h5 mb-0">Autres images de cet objet</h2>
    <span class="text-muted small">{len(items)} image(s)</span>
  </div>
  <div class="row g-3">
    {gallery_cards_html}
  </div>
</main>

{astro_modal}
        {star_modal}

<script src="{BOOTSTRAP_CDN_JS}"></script>
</body>
</html>
"""


def build_app_js() -> str:
    return r"""
function uniq(arr) { return [...new Set(arr)].sort(); }

function fillSelect(select, values) {
  for (const v of values) {
    const opt = document.createElement('option');
    opt.value = v;
    opt.textContent = v;
    select.appendChild(opt);
  }
}

function matches(item, q) {
  if (!q) return true;
  const parts = [
    item.name, item.description, item.objectName, item.catalog, item.objectType,
    item.filter,
    item.messier || "", item.ngc || "", item.constellation || "",
    String(item.magnitude ?? ""), String(item.distance_ly ?? ""), item.size || "",
    ...(item.tags_fr || []), ...(item.tags_en || []),
    ...(item.keywords_fr || []), ...(item.keywords_en || []),
  ];
  const hay = parts.join(' ').toLowerCase();
  return hay.includes(q.toLowerCase());
}

function buildCard(item) {
  const col = document.createElement('div');
  col.className = 'col-12 col-sm-6 col-lg-4';

  const tags = [...new Set([...(item.tags_fr||[]), ...(item.tags_en||[])])]
    .slice(0,6)
    .map(t => `<span class="badge text-bg-secondary me-1 mb-1">${t}</span>`)
    .join('');

  col.innerHTML = `
    <div class="card h-100 shadow-sm">
      <a href="${item.objectPage || item.contentUrl}">
        <img src="${item.thumbnailUrl}" class="card-img-top" alt="${item.alt}">
      </a>
      <div class="card-body">
        <div class="fw-semibold">${item.objectName || item.name}</div>
        <div class="text-muted small">${item.dateCreated || ''}</div>
        <div class="small mt-2">
          <span class="badge text-bg-secondary">${item.objectType || ''}</span>
          <span class="badge text-bg-secondary">${item.catalog || ''}</span>
          <span class="badge text-bg-secondary">${item.filter || ''}</span>
        </div>
        <div class="mt-2">${tags}</div>
      </div>
    </div>
  `;
  return col;
}

function render(data) {
  const grid = document.getElementById('grid');
  const stats = document.getElementById('stats');
  const q = document.getElementById('searchInput').value.trim();
  const type = document.getElementById('typeSelect').value;
  const catalog = document.getElementById('catalogSelect').value;

  const filtered = data.filter(it => {
    if (type && it.objectType !== type) return false;
    if (catalog && it.catalog !== catalog) return false;
    return matches(it, q);
  });

  grid.innerHTML = '';
  for (const item of filtered) grid.appendChild(buildCard(item));
  stats.textContent = `${filtered.length} image(s) affichée(s) / ${data.length} au total`;
}

(function main() {
  let data;
  try {
    data = JSON.parse(document.getElementById('images-data').textContent);
  } catch (e) {
    console.error("Erreur chargement données:", e);
    document.getElementById('stats').textContent = "Impossible de charger les données.";
    return;
  }

  data.sort((a, b) => (b.dateCreatedISO || "").localeCompare(a.dateCreatedISO || ""));

  const typeSelect = document.getElementById('typeSelect');
  const catalogSelect = document.getElementById('catalogSelect');

  fillSelect(typeSelect, uniq(data.map(d => d.objectType).filter(Boolean)));
  fillSelect(catalogSelect, uniq(data.map(d => d.catalog).filter(Boolean)));

  document.getElementById('searchInput').addEventListener('input', () => render(data));
  typeSelect.addEventListener('change', () => render(data));
  catalogSelect.addEventListener('change', () => render(data));

  render(data);
})();
"""


def build_styles_css() -> str:
    return """
.card-img-top { object-fit: cover; height: 240px; }

.object-hero {
  max-height: 70vh;
  object-fit: cover;
}

.astro-preview {
  max-height: 360px;
  object-fit: contain;
}
"""


# ------------------------------------------------------------
# Sitemap
# ------------------------------------------------------------
def write_sitemap(site_dir: Path, urls: list[str], base_url: str):
    urlset = ET.Element("urlset", xmlns="http://www.sitemaps.org/schemas/sitemap/0.9")
    for u in urls:
        url = ET.SubElement(urlset, "url")
        ET.SubElement(url, "loc").text = base_url.rstrip("/") + "/" + u.lstrip("/")
    tree = ET.ElementTree(urlset)
    tree.write(site_dir / "sitemap.xml", encoding="utf-8", xml_declaration=True)


# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
def main():
    root = Path(os.getcwd())
    out = root / "site"

    script_dir = Path(__file__).resolve().parent

    # Messier catalog
    messier_xlsx = find_messier_xlsx(script_dir, root)
    messier_db = {}
    if messier_xlsx:
        try:
            messier_db = load_messier_catalog(messier_xlsx)
            print(f"[INFO] Catalogue Messier chargé: {messier_xlsx} ({len(messier_db)} entrées)")
        except Exception as e:
            print(f"[WARN] Lecture catalogue Messier impossible: {messier_xlsx} ({e})")
            messier_db = {}
    else:
        print(f"[INFO] Catalogue Messier introuvable (attendu: {MESSIER_XLSX_NAME} près du script).")


    # Catalogue d’objets divers (pour labels de la carte): objets <= mag 6
    diverse_xlsx = find_diverse_xlsx(script_dir, root)
    diverse_catalog = []
    if diverse_xlsx:
        try:
            diverse_catalog = load_diverse_catalog(diverse_xlsx)
            print(f"[INFO] Catalogue objets divers chargé: {diverse_xlsx} ({len(diverse_catalog)} entrées)")
        except Exception as e:
            print(f"[WARN] Lecture catalogue objets divers impossible: {diverse_xlsx} ({e})")
            diverse_catalog = []
    else:
        print(f"[INFO] Catalogue objets divers introuvable (attendu: {DIVERSE_XLSX_NAME} près du script).")

    diverse_mag_limit = float(os.environ.get("GNU_ASTRO_GALERY_DIVERSE_MAG_LIMIT", str(DIVERSE_LABEL_MAG_LIMIT_DEFAULT)))
    NOVA_API_KEY = os.environ.get("NOVA_ASTROMETRY_API_KEY", "").strip()
    if not NOVA_API_KEY:
        print("[INFO] NOVA_ASTROMETRY_API_KEY non défini -> pas d'astrométrie (plate solve)")

    # SIMBAD cache
    cache = load_cache(root / CACHE_PATH)

    # Astrometry persistent cache
    ASTRO_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    astro_cache = load_json(ASTRO_CACHE_INDEX)

    # Rebuild site fresh
    if out.exists():
        shutil.rmtree(out)
    (out / "assets/css").mkdir(parents=True, exist_ok=True)
    (out / "assets/js").mkdir(parents=True, exist_ok=True)
    (out / "data").mkdir(parents=True, exist_ok=True)
    (out / "gallery").mkdir(parents=True, exist_ok=True)
    (out / "data/img").mkdir(parents=True, exist_ok=True)
    (out / "astrometry").mkdir(parents=True, exist_ok=True)
    (out / "data/solved").mkdir(parents=True, exist_ok=True)

    jpgs = find_final_jpgs(root)
    if not jpgs:
        print("Aucun JPG final trouvé (hors dossiers *_sub/*-sub et hors fichiers *_thn.jpg).")
        return

    total = len(jpgs)
    processed = 0
    print(f"🔎 {total} image(s) à traiter (hors *_sub/*-sub et hors *_thn.jpg)...")

    items = []
    object_groups = {}

    # Nova session
    nova_session = None
    if NOVA_API_KEY:
        try:
            nova_session = nova_login(NOVA_API_KEY)
            print("[INFO] Nova: session OK")
        except Exception as e:
            print(f"[WARN] Login Nova impossible: {e}")
            nova_session = None

    # Pass 1: build items
    for jpg_path in jpgs:
        processed += 1
        pct = (processed / total) * 100.0
        print(f"🛠️  Scan+tags: {processed}/{total} ({pct:5.1f}%)", end="\r")

        obs_dir = jpg_path.parent
        fits_path = find_stacked_fits_in_dir(obs_dir)
        meta = extract_fits_metadata(fits_path) if fits_path else {}

        object_name = meta.get("object") or obs_dir.name
        if object_name.strip().lower() in ("unknown object", "unknown", ""):
            object_name = obs_dir.name

        dt = parse_date(meta.get("date_obs", ""))
        date_created_human = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        date_created_iso = dt.isoformat() if dt else ""

        catalog = infer_catalog(object_name)
        obj_type = infer_object_type_basic(object_name)

        # Copy image
        rel_img_name = f"{slugify(object_name)}-{jpg_path.name}"
        rel_img = Path("data/img") / rel_img_name
        shutil.copy2(jpg_path, out / rel_img)

        # Thumbnail if exists
        thn_guess = jpg_path.with_name(jpg_path.stem + "_thn.jpg")
        rel_thn = rel_img
        if thn_guess.exists() and (not is_in_sub_folder(thn_guess)):
            rel_thn_name = f"{slugify(object_name)}-{thn_guess.name}"
            rel_thn = Path("data/img") / rel_thn_name
            shutil.copy2(thn_guess, out / rel_thn)

        # SIMBAD: use folder name (not filename, not FITS OBJECT)
        simbad_ident = simbad_ident_from_dir(obs_dir)
        enrich = enrich_tags(simbad_ident, cache)

        tags_fr = uniq_preserve(enrich.get("tags_fr", []))
        tags_en = uniq_preserve(enrich.get("tags_en", []))

        # Improve object type from tags
        if "planetary nebula" in tags_en:
            obj_type = "Planetary Nebula"
        elif "globular cluster" in tags_en:
            obj_type = "Globular Cluster"
        elif "open cluster" in tags_en:
            obj_type = "Open Cluster"
        elif "nebula" in tags_en or "HII region" in tags_en:
            obj_type = "Nebula"
        elif "spiral galaxy" in tags_en:
            obj_type = "Spiral Galaxy"
        elif "galaxy" in tags_en:
            obj_type = "Galaxy"
        elif "star" in tags_en:
            obj_type = "Star"
        elif "planet" in tags_en:
            obj_type = "Planet"

        # Messier enrichment from XLSX
        messier_id = normalize_messier_id(obs_dir.name) or normalize_messier_id(object_name)
        messier_info = messier_db.get(messier_id.upper()) if (messier_id and messier_db) else None

        messier_fields = {
            "messier": "",
            "ngc": "",
            "constellation": "",
            "magnitude": None,
            "size": "",
            "distance_ly": None,
            "messier_type": "",
        }

        if messier_id and messier_info:
            messier_fields["messier"] = messier_id
            messier_fields["ngc"] = messier_info.get("ngc_id") or ""
            messier_fields["constellation"] = messier_info.get("constellation") or ""
            messier_fields["magnitude"] = messier_info.get("mag")
            messier_fields["size"] = messier_info.get("size") or ""
            messier_fields["distance_ly"] = messier_info.get("distance_ly")
            messier_fields["messier_type"] = messier_info.get("type") or ""

            if obj_type in ("Other", "", None) and messier_fields["messier_type"]:
                obj_type = messier_fields["messier_type"]

            if messier_fields["messier_type"]:
                tags_fr = uniq_preserve(tags_fr + [messier_fields["messier_type"]])

        keywords_fr = uniq_preserve([object_name, catalog, "Seestar S50", "astrophotographie"] + tags_fr)
        keywords_en = uniq_preserve([object_name, catalog, "Seestar S50", "astrophotography"] + tags_en)

        if messier_fields["messier"]:
            keywords_fr = uniq_preserve(keywords_fr + [messier_fields["messier"], messier_fields["ngc"], messier_fields["constellation"]])
            keywords_en = uniq_preserve(keywords_en + [messier_fields["messier"], messier_fields["ngc"], messier_fields["constellation"]])

        tags_short_fr = ", ".join(tags_fr[:3]) if tags_fr else ""
        tags_short_en = ", ".join(tags_en[:3]) if tags_en else ""
        desc = f"Astrophotographie {object_name} (Seestar S50). {tags_short_fr} / {tags_short_en}".strip()

        alt = f"{object_name} — {', '.join(uniq_preserve(tags_fr[:2] + tags_en[:2]))} — {catalog} — Seestar S50".strip(" —")

        obj_slug = slugify(object_name)
        object_page = f"gallery/{obj_slug}.html"

        content_abs = BASE_URL.rstrip("/") + "/" + rel_img.as_posix()
        thumb_abs = BASE_URL.rstrip("/") + "/" + rel_thn.as_posix()

        item = {
            "name": jpg_path.name,
            "objectName": object_name,
            "catalog": catalog,
            "objectType": obj_type,
            "dateCreated": date_created_human,
            "dateCreatedISO": date_created_iso,
            "filter": meta.get("filter", ""),
            "exptime": meta.get("exptime", ""),
            "telescope": meta.get("telescope", ""),
            "instrument": meta.get("instrument", ""),
            "ra": meta.get("ra", ""),
            "dec": meta.get("dec", ""),

            "tags_fr": tags_fr,
            "tags_en": tags_en,
            "keywords_fr": keywords_fr,
            "keywords_en": keywords_en,

            # SIMBAD fields to display on object page
            "simbad_ident": enrich.get("ident", ""),
            "simbad_main_id": enrich.get("main_id", ""),
            "simbad_otype": enrich.get("otype", ""),
            "simbad_otype_txt": enrich.get("otype_txt", ""),
            "simbad_source": enrich.get("source", ""),

            # Messier fields
            "messier": messier_fields["messier"],
            "ngc": messier_fields["ngc"],
            "constellation": messier_fields["constellation"],
            "magnitude": messier_fields["magnitude"],
            "size": messier_fields["size"],
            "distance_ly": messier_fields["distance_ly"],
            "messier_type": messier_fields["messier_type"],

            "description": desc,
            "alt": alt,

            "contentUrl": rel_img.as_posix(),
            "thumbnailUrl": rel_thn.as_posix(),
            "contentUrlAbs": content_abs,
            "thumbnailUrlAbs": thumb_abs,

            "objectPage": object_page,
            "astrometryUrl": "",

            "_fitsPath": str(fits_path) if fits_path else "",
            "_jpgPath": str(jpg_path),
            "_jpgStem": jpg_path.stem,
        }

        items.append(item)
        object_groups.setdefault(object_name, []).append(item)

    save_cache(root / CACHE_PATH, cache)
    print("\n✅ Tags + Messier: terminé (cache SIMBAD mis à jour).")

    items.sort(key=lambda x: x.get("dateCreatedISO", ""), reverse=True)

    # Pass 2: astrometry (optional) + persistent cache
    if nova_session:
        if ASTROMETRY_MODE == "all":
            to_solve = [it for it in items if it.get("_fitsPath")]
        else:
            to_solve = []
            for obj, group in object_groups.items():
                group.sort(key=lambda x: x.get("dateCreatedISO", ""), reverse=True)
                if group and group[0].get("_fitsPath"):
                    to_solve.append(group[0])

        print(f"[INFO] Astrométrie mode={ASTROMETRY_MODE} -> {len(to_solve)} solve(s)")

        done = 0
        for it in to_solve:
            done += 1
            obj = it["objectName"]
            fits_path = Path(it["_fitsPath"]) if it.get("_fitsPath") else None
            jpg_path = Path(it["_jpgPath"])
            jpg_stem = it["_jpgStem"]
            pct = (done / max(1, len(to_solve))) * 100.0
            print(f"🌐 Astrométrie: {done}/{len(to_solve)} ({pct:5.1f}%) — {obj}", end="\r")

            try:
                if not fits_path or not fits_path.exists():
                    print(f"\n[WARN] Pas de FITS local pour {obj} -> astrométrie skip")
                    continue

                # ---------- Persistent cache check ----------
                src_path = fits_path if fits_path.exists() else jpg_path
                src_fp = file_fingerprint(src_path)

                cache_key = f"{slugify(obj)}-{jpg_stem}"
                cached = astro_cache.get(cache_key)

                cached_png = ASTRO_CACHE_DIR / f"{cache_key}-astrometry.png"
                cached_wcs = ASTRO_CACHE_DIR / f"{cache_key}-wcs.fits"

                if cached and cached.get("src_fp") == src_fp and cached_png.exists() and cached_wcs.exists():
                    astro_rel = Path("astrometry") / cached_png.name
                    wcs_rel = Path("data/solved") / cached_wcs.name

                    shutil.copy2(cached_png, out / astro_rel)
                    shutil.copy2(cached_wcs, out / wcs_rel)

                    it["astrometryUrl"] = astro_rel.as_posix()
                    print(f"\n[CACHE] Astrométrie réutilisée pour {obj}: {cached_png.name}")
                    continue
                # -------------------------------------------

                wcs_fits = out / "data/solved" / f"{cache_key}-wcs.fits"

                # Run solve if we don't already have WCS for this run
                if not wcs_fits.exists():
                    scale = estimate_scale_arcsec_per_pix(fits_path)
                    subid = nova_upload_fits(nova_session, fits_path, scale_arcsec_per_pix=scale)
                    jobid = nova_poll_submission(subid, wait_s=5, timeout_s=600)
                    ok = nova_poll_job_solved(jobid, wait_s=5, timeout_s=900)
                    if not ok:
                        print(f"\n[WARN] Plate-solve échoué (job failure) pour {obj}")
                        continue

                    ok_dl = nova_download_wcs_header_only(jobid, wcs_fits)
                    if not ok_dl:
                        print(f"\n[WARN] Téléchargement WCS header-only échoué pour {obj}")
                        continue

                wcs_header = load_wcs_header_only(wcs_fits)
                if wcs_header is None:
                    continue
                # Centre du champ (WCS) pour usage local (carte stellaire)
                ra_c, dec_c = wcs_center_from_header(wcs_header)
                if ra_c is not None and dec_c is not None:
                    it["wcsCenterRaDeg"] = ra_c
                    it["wcsCenterDecDeg"] = dec_c

                img = read_best_image_from_fits(fits_path)
                if img is None:
                    print(f"\n[WARN] Image FITS invalide: aucun HDU 2D dans {fits_path.name} -> fallback JPG")
                    img = read_image_from_jpg(jpg_path)

                if img is None:
                    print(f"\n[WARN] Aucun pixel image disponible (FITS+JPG) pour {obj}")
                    continue

                astro_name = f"{cache_key}-astrometry.png"
                astro_rel = Path("astrometry") / astro_name

                ok_png = make_astrometry_png_from_image_and_wcs(
                    image_array_2d=img,
                    wcs_header=wcs_header,
                    out_png=out / astro_rel,
                    title=obj
                )

                if ok_png:
                    it["astrometryUrl"] = astro_rel.as_posix()

                # Carte stellaire (cache persistant + copie dans /site)
                try:
                    ra_c = it.get("wcsCenterRaDeg", None)
                    dec_c = it.get("wcsCenterDecDeg", None)

                    if ra_c is None or dec_c is None:
                        # fallback Messier si disponible (si vous le mappez plus tard)
                        ra_c = it.get("messier_ra_deg", None)
                        dec_c = it.get("messier_dec_deg", None)

                    if ra_c is not None and dec_c is not None:
                        STAR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
                        star_cache = load_star_cache()
                        star_key = f"{obj.upper()}|{float(ra_c):.6f}|{float(dec_c):.6f}|30"

                        if star_key in star_cache and Path(star_cache[star_key]).exists():
                            star_png_cache = Path(star_cache[star_key])
                        else:
                            star_png_cache = STAR_CACHE_DIR / f"{slugify(obj)}_{abs(int(float(ra_c)*1000))}_{abs(int(float(dec_c)*1000))}_30.png"
                            ok_star = make_finder_chart_png(float(ra_c), float(dec_c), star_png_cache, fov_arcmin=ATLAS_FOV_ARCMIN, inner_fov_arcmin=30.0, title=obj, diverse_catalog=diverse_catalog, diverse_mag_limit=diverse_mag_limit)
                            if ok_star:
                                star_cache[star_key] = str(star_png_cache)
                                save_star_cache(star_cache)

                        if star_png_cache.exists():
                            (out / "starcharts").mkdir(parents=True, exist_ok=True)
                            star_name = f"{slugify(obj)}-finder.png"
                            dest_rel = Path("starcharts") / star_name
                            shutil.copy2(star_png_cache, out / dest_rel)
                            it["starChartUrl"] = dest_rel.as_posix()
                except Exception as e:
                    print(f"[WARN] Carte stellaire: erreur inattendue: {e}")

                    # ---------- Save to persistent cache ----------
                    shutil.copy2(wcs_fits, cached_wcs)
                    shutil.copy2(out / astro_rel, cached_png)
                    astro_cache[cache_key] = {
                        "src_fp": src_fp,
                        "object": obj,
                        "updated": datetime.now().isoformat(timespec="seconds")
                    }
                    save_json(ASTRO_CACHE_INDEX, astro_cache)
                    # ---------------------------------------------

                else:
                    print(f"\n[WARN] PNG astrométrie non généré: {obj}")

            except Exception as e:
                print(f"\n[WARN] Astrométrie échouée pour {obj}: {e}")

        print("\n✅ Astrométrie: terminé.")
    else:
        print("[INFO] Astrométrie non exécutée (pas de session Nova).")

    # Remove internal fields
    for it in items:
        it.pop("_fitsPath", None)
        it.pop("_jpgPath", None)
        it.pop("_jpgStem", None)

    print("🧩 Écriture des fichiers du site...")

    (out / "data/images.json").write_text(
        json.dumps(items, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    og_img = items[0]["thumbnailUrlAbs"] if items else (BASE_URL.rstrip("/") + "/")
    og_index = og_meta(SITE_TITLE, "Galerie astrophotographie — Seestar S50", og_img, BASE_URL.rstrip("/") + "/index.html")

    index_html = build_index_html(SITE_TITLE, og_index)
    index_html = index_html.replace(
        '<script id="images-data" type="application/json"></script>',
        '<script id="images-data" type="application/json">\n' +
        json.dumps(items, ensure_ascii=False) +
        '\n</script>'
    )
    (out / "index.html").write_text(index_html, encoding="utf-8")

    (out / "assets/js/app.js").write_text(build_app_js(), encoding="utf-8")
    (out / "assets/css/styles.css").write_text(build_styles_css(), encoding="utf-8")

    # Object pages
    object_urls = []
    for obj_name, group_items in object_groups.items():
        group_items.sort(key=lambda x: x.get("dateCreatedISO", ""), reverse=True)

        top = group_items[0]
        page_rel = f"gallery/{slugify(obj_name)}.html"
        page_url_abs = BASE_URL.rstrip("/") + "/" + page_rel

        og_obj = og_meta(
            f"{obj_name} — {SITE_TITLE}",
            top.get("description", f"Astrophotographie: {obj_name}"),
            top["thumbnailUrlAbs"],
            page_url_abs
        )

        jsonld = image_jsonld(top, page_url_abs)

        page = build_object_page_html(
            SITE_TITLE,
            obj_name,
            json.dumps(jsonld, ensure_ascii=False, indent=2),
            og_obj,
            group_items
        )

        page_path = out / "gallery" / f"{slugify(obj_name)}.html"
        page_path.write_text(page, encoding="utf-8")
        object_urls.append(page_rel)

    (out / "robots.txt").write_text(
        "User-agent: *\nAllow: /\nSitemap: sitemap.xml\n",
        encoding="utf-8"
    )

    write_sitemap(out, ["index.html"] + object_urls, BASE_URL)

    print(f"✅ Galerie générée dans: {out}")
    print(f"📌 Cache SIMBAD: {root / CACHE_PATH}")
    print(f"📌 Cache astrométrie: {ASTRO_CACHE_DIR} (index: {ASTRO_CACHE_INDEX})")
    if BASE_URL.startswith("https://example.com"):
        print("⚠️  Mets BASE_URL sur ton URL réelle si tu publies (sinon OG/sitemap ont une URL fictive).")


if __name__ == "__main__":
    main()